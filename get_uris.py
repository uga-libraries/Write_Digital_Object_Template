import json
import os
import platform
import subprocess
import sys
from pathlib import Path
from asnake.client import ASnakeClient
from asnake.client.web_client import ASnakeAuthError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import PySimpleGUI as psg
import requests


def gui():
    """
    The main GUI for the application, sets user input for filepaths and displays output console.

    :arg:
        None

    :return:
        None
    """
    defaults = psg.UserSettings()
    close_program, client, repositories = get_aspace_log(defaults)
    if close_program is True:
        sys.exit()
    main_layout = [[psg.Text("Choose your repository:", font=("Roboto", 12))],
                   [psg.DropDown(list(repositories.keys()), readonly=True,
                                 default_value=defaults["repo_default"], key="_REPO_SELECT_",),
                    psg.Button(" SAVE ", key="_SAVE_REPO_")],
                   [psg.FileBrowse(' Select Digital Objects File ',
                                   file_types=(("Excel Files", "*.xlsx"),),),
                    psg.InputText(default_text=defaults['_DO_FILE_'], key='_DO_FILE_')],
                   [psg.FileBrowse(' Select Template ', file_types=(("Excel Files", "*.xlsx"),),),
                    psg.InputText(default_text=defaults['_DOTEMP_FILE_'], key='_DOTEMP_FILE_')],
                   [psg.Button(' START ', key='_WRITE_DOS_', disabled=False)],
                   [psg.Output(size=(80, 18), key="_output_")],
                   [psg.Button(" Open DO Template File ", key="_OPEN_DOTEMP_")]]
    window = psg.Window('Write Digital Objects to Template', layout=main_layout)
    while True:
        event, values = window.read()
        if event in (psg.WINDOW_CLOSED, 'Exit'):
            break
        if event == '_WRITE_DOS_':
            if not values['_DO_FILE_']:
                psg.popup_error('ERROR\nPlease select a digital objects file',
                                font=("Roboto", 14), keep_on_top=True)
            elif not values['_DOTEMP_FILE_']:
                psg.popup_error('ERROR\nPlease select a digital object template',
                                font=("Roboto", 14), keep_on_top=True)
            else:
                defaults['_DO_FILE_'] = values['_DO_FILE_']
                defaults['_DOTEMP_FILE_'] = values['_DOTEMP_FILE_']
                write_digobjs(values['_DO_FILE_'], values['_DOTEMP_FILE_'], client,
                              repositories[values["_REPO_SELECT_"]], window)
        if event == "_SAVE_REPO_":
            defaults["repo_default"] = values["_REPO_SELECT_"]
        if event == "_OPEN_DOTEMP_":
            if not defaults["_DOTEMP_FILE_"]:
                filepath_eads = str(Path.cwd())
                open_file(filepath_eads)
            else:
                filepath_eads = str(Path(defaults["_DOTEMP_FILE_"]))
                open_file(filepath_eads)


def write_digobjs(digobj_file, dotemp_file, client, repo, gui_window):
    """
    Fetches data from digital object spreadsheet and ArchivesSpace and
    writes data to digital object import template.

    To import digital object data from another spreadsheet, the following
    columns must be present and in this order:
        digital_object_id - Column 0; digital_object_title - Column 2;
        file_version_file_uri - Column 3; date_1_expression - Column 5;
        digital_object_publish - Column 8
    To match the appropriate archival object, the digital object title
    and date must match the archival object title and date.

    Args:
        :param (str) digobj_file: Filepath for the digital object spreadsheet.
        :param (str) dotemp_file: Filepath for the ArchivesSpace digital
        object importer spreadsheet.
        :param (ASnakeClient) client: ArchivesSpace ASnake client for
        accessing and connecting to the API.
        :param (int) repo: ArchivesSpace URI number for user selected
        repository.
        :param (PySimpleGUI Window) gui_window: The PySimpleGUI window used
        for threading.

    :return:
        error (str): Errors caught while running function
    """
    gui_window[f'{"_WRITE_DOS_"}'].update(disabled=True)
    digobj_wb = load_workbook(digobj_file)
    digobj_sheet = digobj_wb.active
    dotemp_wb = load_workbook(dotemp_file)
    dotemp_sheet = dotemp_wb.active
    write_row_index = 6
    total_digobjs = 0
    digobj_columns = {0: "digital_object_id", 2: "digital_object_title",
                      3: "file_version_file_uri", 5: "date_1_expression",
                      8: "digital_object_publish"}
    sheet_columns = {}
    sheet_colnum = 0
    for col in digobj_sheet.iter_cols(max_col=digobj_sheet.max_column,
                                      values_only=True):
        sheet_columns[sheet_colnum] = col[0]
        sheet_colnum += 1
    for col_key, col_value in digobj_columns.items():
        if sheet_columns[col_key] != col_value:
            psg.popup_error("ERROR:\nDigital Object file columns do not match!\n\n"
                            "Check to make sure you entered the correct file")
            error = "ERROR: Digital Object file columns do not match!\n\n" \
                    "Check to make sure you entered the correct file"
            close_wbs(digobj_wb, dotemp_wb)
            gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
            return error
    errors = []
    for row in digobj_sheet.iter_rows(min_row=2, values_only=True):
        write_row_index += 1
        total_digobjs += 1
        digobj_id = row[0]
        digobj_title = row[2]
        digobj_url = row[3]
        digobj_date = row[5]
        digobj_publish = row[8]
        archobj_uri, resource_uri = get_results(client,
                                                repo,
                                                digobj_title,
                                                digobj_date)
        if archobj_uri is None and resource_uri is None:
            archobj_uri = "!!ERROR!!"
            resource_uri = "!!ERROR!!"
            errors.append(f'{digobj_title}, {digobj_date}')
            for cell in dotemp_sheet[f'{write_row_index}:{write_row_index}']:
                cell.fill = PatternFill(start_color='FFFF0000',
                                        end_color='FFFF0000',
                                        fill_type='solid')
        write_obj_error = write_digobj(resource_uri, archobj_uri, digobj_id,
                                       digobj_title, digobj_publish, digobj_url,
                                       dotemp_sheet, write_row_index, dotemp_wb,
                                       dotemp_file, gui_window)
        if write_obj_error is not None:
            print(write_obj_error)
            close_wbs(digobj_wb, dotemp_wb)
        else:
            print(f'{digobj_title}, {digobj_date}')
    close_wbs(digobj_wb, dotemp_wb)
    print(f'\n{"*" * 112}\n{" " * 40}Finished writing {total_digobjs} to '
          f'{dotemp_sheet}\n{"*" * 112}')
    if errors:
        error_message = "\nERROR: Could not find any records with the " \
                        "following titles:\n"
        for error in errors:
            error_message += "\n" + error + "\n"
        print(error_message)
    gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
    return errors


def close_wbs(digobj_wb, dotemp_wb):
    digobj_wb.close()
    dotemp_wb.close()


def get_results(client, repo, digobj_title, digobj_date):
    """
    Searches for archival objects in ArchivesSpace and returns results.

    To search for archival objects to link digital objects to, the
    function tries searching for the "title, date" of the digital
    object supplied by the digital object spreadsheet and if it
    finds an archival object with a matching "title, date", then it
    grabs the URI for both the archival object and resource and
    returns them. If more than 1 archival object is found, a popup
    is generated asking the user which archival object they should
    choose. If no archival object is found, it returns None and
    prints an error message.

    Args:
        :param (ASnakeClient) client: ArchivesSpace ASnake client
        for accessing and connecting to the API.
        :param (int) repo: ArchivesSpace URI number for user selected
        repository.
        :param (str) digobj_title: Title of the digital object.
        :param (str) digobj_date: Date of the digital object.

    :return:
        archobj_uri (str): URI to matched archival object or NONE
        resource_uri (str): URI to resource for matched archival
        object or NONE
    """
    archobj_uri = None
    resource_uri = None
    search_archobjs = client.get_paged(f"/repositories/{repo}/search",
                                       params={"q": f'title:"{digobj_title}, '
                                                    f'{digobj_date}"',
                                               "type": ['archival_object']})
    search_results = []
    for results in search_archobjs:
        search_results.append(results)
    if len(search_results) > 1:
        search_options = []
        for result in search_results:
            result_container_uri = result["top_container_uri_u_sstr"][0]
            top_container_json = client.get(result_container_uri).json()
            box_coll_info = top_container_json["long_display_string"]
            box_coll_list = box_coll_info.split(",")
            result_child = result["child_container_u_sstr"][0]
            result_option = f'{result["title"]}; ' \
                            f'{box_coll_list[0]}; ' \
                            f'{result_child}; ' \
                            f'{box_coll_list[1]}'
            search_options.append(result_option)
        multresults_layout = [[psg.Text(f'\n\nFound multiple options for'
                                        f'\n{digobj_title}, {digobj_date}\n\n'
                                        f'Choose one of the following:\n')],
                              [psg.Listbox(search_options, size=(120, 5),
                                           key="_ARCHOBJ_FILE_")],
                              [psg.Button(" SELECT ", key="_SELECT_ARCHOBJ_")]]
        multresults_window = psg.Window("Multiple Results for Archival Object",
                                        multresults_layout)
        selection = True
        while selection is True:
            multresults_event, multresults_values = multresults_window.Read()
            if multresults_event == "_SELECT_ARCHOBJ_":
                result_title = multresults_values["_ARCHOBJ_FILE_"][0].split(";")[0]
                for result in search_results:
                    if result["title"] == result_title:
                        archobj_uri = result["uri"]
                        resource_uri = result["resource"]
                        selection = False
                        multresults_window.close()
                        break
    elif len(search_results) == 0:
        print(f'\nERROR: No results found for:\n{digobj_title}, {digobj_date}\n')
        return archobj_uri, resource_uri
    else:
        for result in search_results:
            archobj_uri = result["uri"]
            resource_uri = result["resource"]
    if archobj_uri is None and resource_uri is None:
        print(f'{digobj_title}, {digobj_date}')
        print(search_results)
    return archobj_uri, resource_uri


def write_digobj(resource_uri, archobj_uri, digobj_id, digobj_title,
                 digobj_publish, digobj_url, dotemp_sheet, write_row_index,
                 dotemp_wb, dotemp_file, gui_window):
    """
    Takes parameters and writes them to the ArchivesSpace digital object
    import spreadsheet template.

    Args:
        :param (str) resource_uri: Resource URI for a matched archival object.
        :param (str) archobj_uri: Matched archival object URI.
        :param (str) digobj_id: Digital object ID provided by digital object
        spreadsheet.
        :param (str) digobj_title: Digital object title provided by the digital
        object spreadsheet.
        :param (bool) digobj_publish: Digital object publish status provided by
        the digital object spreadsheet.
        :param (str) digobj_url: Digital object file version URL provided by the
        digital object spreadsheet.
        :param (openpyxl Worksheet) dotemp_sheet: openpyxl worksheet for
        ArchivesSpace digital object import template.
        :param (int) write_row_index: Number to keep track of which row to write
        to in spreadsheet.
        :param (openpyxl Workbook) dotemp_wb: openpyxl workbook for ArchivesSpace
        digital object import template.
        :param (str) dotemp_file: Filepath of the ArchivesSpace digital object
        import template.
        :param (PySimpleGUI Window) gui_window: The PySimpleGUI window used for
        threading.

    :return:
        Returns an error message (str) if an error occurred. Otherwise returns NONE.
    """
    column_map = {4: resource_uri,
                  6: archobj_uri,
                  7: digobj_id,
                  8: digobj_title,
                  9: digobj_publish,
                  10: digobj_url}
    for column_num, column_value in column_map.items():
        dotemp_sheet.cell(row=write_row_index, column=column_num).value = column_value
    write_row_index += 1
    try:
        dotemp_wb.save(dotemp_file)
        return None
    except Exception as save_exception:
        error = f'\n\nFailed opening {dotemp_file}. ' \
                f'Please close the record before trying again.\nError: {save_exception}'
        gui_window[f'{"_WRITE_DOS_"}'].update(disabled=False)
        return error


def get_aspace_log(defaults):
    """
    Gets a user's ArchiveSpace credentials.

    There are 3 components to it, the setup code, correct_creds while loop,
    and the window_asplog_active while loop. It uses ASnake.client to
    authenticate and stay connected to ArchivesSpace. Documentation for ASnake
    can be found here:
    https://archivesspace-labs.github.io/ArchivesSnake/html/index.html

    Args:
        :param defaults: contains data from PySimpleGUI's defaults.json file.

    :return:
        close_program (bool): if a user exits the popup, this will return true
        and end run_gui().
        connect_client (ASnakeClient): the ArchivesSpace ASnake client for
        accessing and connecting to the API.
        repositories (dict): Names and URI number for all repositories in
        ArchivesSpace instance.
    """
    connect_client = None
    repositories = {}
    save_button_asp = " Save and Continue "
    window_asplog_active = True
    correct_creds = False
    close_program = False
    while correct_creds is False:
        asplog_col1 = [[psg.Text("ArchivesSpace username:",
                                 font=("Roboto", 11))],
                       [psg.Text("ArchivesSpace password:",
                                 font=("Roboto", 11))],
                       [psg.Text("ArchivesSpace API URL:",
                                 font=("Roboto", 11))]]
        asplog_col2 = [[psg.InputText(focus=True,
                                      key="_ASPACE_UNAME_")],
                       [psg.InputText(password_char='*',
                                      key="_ASPACE_PWORD_")],
                       [psg.InputText(defaults["as_api"],
                                      key="_ASPACE_API_")]]
        layout_asplog = [
            [psg.Column(asplog_col1,
                        key="_ASPLOG_COL1_",
                        visible=True),
             psg.Column(asplog_col2,
                        key="_ASPLOG_COL2_",
                        visible=True)],
            [psg.Button(save_button_asp,
                        bind_return_key=True,
                        key="_SAVE_CLOSE_LOGIN_")]
        ]
        window_login = psg.Window("ArchivesSpace Login Credentials",
                                  layout_asplog)
        while window_asplog_active is True:
            event_log, values_log = window_login.Read()
            if event_log == "_SAVE_CLOSE_LOGIN_":
                connect_client = ASnakeClient(baseurl=values_log["_ASPACE_API_"],
                                              username=values_log["_ASPACE_UNAME_"],
                                              password=values_log["_ASPACE_PWORD_"])
                try:
                    requests.get(values_log["_ASPACE_API_"])
                except Exception as api_error:
                    psg.Popup("Your API credentials were entered incorrectly.\n"
                              "Please try again.\n\n" + api_error.__str__())
                else:
                    try:
                        connect_client.authorize()
                    except ASnakeAuthError as connection_error:
                        error_message = ""
                        if ":" in str(connection_error):
                            error_divided = str(connection_error).split(":")
                            for line in error_divided:
                                error_message += line + "\n"
                        else:
                            error_message = str(connection_error)
                        psg.Popup("Your username and/or password were entered\n "
                                  "incorrectly. Please try again.\n\n" +
                                  error_message)
                    else:
                        defaults["as_api"] = values_log["_ASPACE_API_"]
                        repo_results = connect_client.get('/repositories')
                        repo_results_dec = json.loads(repo_results.content.decode())
                        for result in repo_results_dec:
                            uri_components = result["uri"].split("/")
                            repositories[result["name"]] = int(uri_components[-1])
                        window_asplog_active = False
                        correct_creds = True
            if event_log is None or event_log == 'Cancel':
                window_login.close()
                window_asplog_active = False
                correct_creds = True
                close_program = True
                break
        window_login.close()
    return close_program, connect_client, repositories


def open_file(filepath):
    """
    Takes a filepath and opens the folder according to Windows, Mac,
    or Linux.

    Args:
        :param (str) filepath: Filepath of the folder/directory a user
        wants to open.

    :return:
        None
    """
    if platform.system() == "Windows":
        os.startfile(filepath)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", filepath])
    else:
        subprocess.Popen(["xdg-open", filepath])


if __name__ == "__main__":
    gui()
